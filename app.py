"""Nuru: user-facing interface. PDFs in (invoices, receipts, or financial
statements), a single Excel workbook out.

Runs the trained pure-NumPy token classifier over the text extracted from
each PDF, decodes the BIO tag sequence into whichever fields are relevant
to that document's kind, and writes one row per document to an .xlsx file.

Different document kinds surface different fields: an invoice reports a
supplier and a tax value, a receipt reports a merchant and a payment
method, a statement reports an account and a closing balance. Which kind a
given file is gets inferred from which entities the model actually found
in it (see infer_document_type below) rather than being fixed in advance,
so the field set genuinely depends on what was uploaded.

Real documents are messy: a page can fail to parse, an amount can hit an
untrained corner of the model. Two things follow from that:
  - Extraction is wrapped in per-page and per-document error boundaries, so
    one bad page or one unreadable PDF never aborts the rest of a bulk run.
  - Every predicted field carries its softmax confidence; anything under
    the threshold is flagged with "Needs Review" and highlighted orange in
    the workbook so a human double-checks it instead of trusting it blind.

Usage:
    python app.py invoice1.pdf receipt2.pdf --output extracted.xlsx
    python app.py documents/*.pdf
    python app.py documents/*.pdf --purge-source   # delete PDFs once captured in the .xlsx
"""

import argparse
import datetime
import glob
import math
import os
import re

import pdfplumber
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import errors
from engine.model import ID2LABEL, TokenClassifier, build_context_windows
from engine.tokenizer import Vocabulary, tokenize

_logger = errors.get_logger()

ENGINE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "engine")
WEIGHTS_PATH = os.path.join(ENGINE_DIR, "weights.npz")
VOCAB_PATH = os.path.join(ENGINE_DIR, "vocab.json")

DEFAULT_CONFIDENCE_THRESHOLD = 0.70

# Underlying entity -> the professional term shown to the user, per document
# kind. Same extraction model, different vocabulary depending on what was
# actually uploaded: a "Total" reads as "Total Amount Due" on an invoice but
# "Amount Paid" on a receipt.
TYPE_FIELD_CONFIG = {
    "Invoice": [
        ("Vendor", "Supplier Name"),
        ("Date", "Transaction Date"),
        ("Total", "Total Amount Due"),
        ("Tax", "Tax Value (VAT/GST)"),
    ],
    "Receipt": [
        ("Merchant", "Merchant"),
        ("Date", "Purchase Date"),
        ("Total", "Amount Paid"),
        ("PaymentMethod", "Payment Method"),
    ],
    "Statement": [
        ("Account", "Account Name"),
        ("Period", "Statement Period"),
        ("Balance", "Closing Balance"),
    ],
}

# Entity-name fields (Vendor/Merchant/Account) are weak signals for type:
# an unfamiliar name's span can get mislabeled as the wrong entity kind
# (an out-of-vocabulary "X & Y" name can read as either a company or a
# merchant), and the model has no way to tell from that name alone. Tax,
# Balance/Period, and PaymentMethod don't have that problem: their VALUES
# only ever appear in one document kind's training data, so if the model
# found one at all, it's much stronger evidence than a name-based guess.
# Strong indicators are checked first and win outright; name-based ones
# are only a tiebreaker when nothing decisive fired.
_STRONG_TYPE_INDICATORS = [
    ("Statement", {"Balance", "Period"}),
    ("Receipt", {"PaymentMethod"}),
    ("Invoice", {"Tax"}),
]
_WEAK_TYPE_INDICATORS = [
    ("Receipt", {"Merchant"}),
    ("Statement", {"Account"}),
    ("Invoice", {"Vendor"}),
]

MAX_FIELDS_PER_TYPE = max(len(cfg) for cfg in TYPE_FIELD_CONFIG.values())

# The same navy/cyan identity as the web UI (webapp.py's BASE_STYLE), so the
# spreadsheet a reviewer downloads looks like it came from the same product,
# not a generic script export.
_NAVY = "16205C"
_NAVY_DARK = "0E1740"
_CYAN_TINT = "E8F9FD"
_LINE = "D7DCE8"
_TEXT = "16203D"
_TEXT_SOFT = "6B7590"
_OK_TEXT, _OK_TINT = "2F6F52", "E9F4EE"
_WARN_TEXT, _WARN_TINT = "A8621A", "FBEDD8"
_ERROR_TEXT, _ERROR_TINT = "A8402F", "FAEAE6"
_RECEIPT_TEXT, _RECEIPT_TINT = "1F6F4A", "E3F6EC"
_STATEMENT_TEXT, _STATEMENT_TINT = "4B3B93", "EFEAFB"
_TYPE_CHIP = {
    "Invoice": (_NAVY, _CYAN_TINT),
    "Receipt": (_RECEIPT_TEXT, _RECEIPT_TINT),
    "Statement": (_STATEMENT_TEXT, _STATEMENT_TINT),
}

HEADER_FILL = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
BAND_FILL = PatternFill(start_color=_CYAN_TINT, end_color=_CYAN_TINT, fill_type="solid")
LOW_CONFIDENCE_FILL = PatternFill(start_color=_WARN_TINT, end_color=_WARN_TINT, fill_type="solid")
ERROR_FILL = PatternFill(start_color=_ERROR_TINT, end_color=_ERROR_TINT, fill_type="solid")
VERIFY_OK_FILL = PatternFill(start_color=_OK_TINT, end_color=_OK_TINT, fill_type="solid")
VERIFY_WARN_FILL = PatternFill(start_color=_WARN_TINT, end_color=_WARN_TINT, fill_type="solid")

_THIN = Side(style="thin", color=_LINE)
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FONT_FAMILY = "Calibri"


def _font(size=11, bold=False, italic=False, color=_TEXT):
    return Font(name=_FONT_FAMILY, size=size, bold=bold, italic=italic, color=color)


_MONEY_RE = re.compile(r"^\(?-?\$?\s?[\d,]+(?:\.\d+)?\)?$")


def _parse_money(value):
    """Recognizes a plain currency-looking string ("$1,234.56", "(200.00)")
    so it can be written as a real Excel number with currency formatting
    instead of inert text, letting a reviewer sum or sort amounts natively
    rather than re-typing them. Returns None for anything else and never
    raises."""
    if not value:
        return None
    text = value.strip()
    if not _MONEY_RE.match(text):
        return None
    negative = text.startswith("(") and text.endswith(")")
    cleaned = text.strip("()").replace("$", "").replace(",", "").strip()
    try:
        number = float(cleaned)
    except ValueError:
        return None
    return -number if negative else number


def _wrapped_row_height(text, width_chars, base_height=18, line_height=14):
    """A column that's allowed to grow without bound for one long note
    would force every other column absurdly wide too. Capping that
    column's width and instead wrapping the text, with the row grown
    tall enough to show every line, is how a real template handles
    unpredictable-length text without asking anyone to resize anything."""
    if not text:
        return base_height
    lines = max(1, math.ceil(len(text) / width_chars))
    return max(base_height, lines * line_height)


TYPE_INFERENCE_CONFIDENCE_FLOOR = 0.60


def infer_document_type(entity_confidences, floor=TYPE_INFERENCE_CONFIDENCE_FLOOR):
    """Pick a document kind from the entities the model found, but only
    the ones it actually found with some conviction, and weighted so an
    ambiguous entity-name guess can't outvote a structurally unambiguous
    field. `entity_confidences` maps entity key -> list of confidences for
    every span of that entity in the document.

    Two real failure modes drove this design, both found by evaluating
    against held-out documents rather than assumed:
      1. An out-of-vocabulary name embeds as <UNK>, and the untrained
         <UNK> embedding produces near-coin-flip predictions; the floor
         below stops a low-confidence guess from counting as evidence.
      2. Even a *confident* wrong guess is possible: "Harrow & Finch Legal
         Services" pattern-matched training's merchant names ("Marlowe &
         Finch Booksellers") well enough to tag as Merchant at 0.95
         confidence, while a correctly-detected Tax field (impossible on
         a receipt or statement in training) sat right next to it,
         unused. Checking the structurally-unambiguous fields
         (Tax/Balance-Period/PaymentMethod) first, before ever consulting
         a name field, fixes this without needing the name guess to stop
         happening at all."""
    confident_entities = {
        key for key, confs in entity_confidences.items()
        if confs and max(confs) >= floor
    }
    for type_name, indicators in _STRONG_TYPE_INDICATORS:
        if confident_entities & indicators:
            return type_name
    for type_name, indicators in _WEAK_TYPE_INDICATORS:
        if confident_entities & indicators:
            return type_name
    return "Invoice"


def _ocr_page(page):
    """Best-effort OCR for a page with no text layer (a scanned or
    photographed document). Degrades to "" and never raises, whether
    pytesseract isn't installed or the Tesseract engine itself isn't on
    this machine; the caller falls back to the normal "no readable text"
    path in that case."""
    try:
        import pytesseract
    except ImportError:
        return ""
    try:
        image = page.to_image(resolution=200).original
        return pytesseract.image_to_string(image) or ""
    except Exception as exc:
        _logger.warning(f"OCR unavailable or failed: {exc}")
        return ""


def extract_text(pdf_path):
    """Extract text page by page.

    A page that fails to parse (corrupt content stream, malformed table,
    unsupported encoding) is skipped rather than raising, so it doesn't take
    the rest of the document down with it. Returns (text, failed_page_numbers, used_ocr).

    If no page yields a real text layer at all (common for a scanned or
    phone-photographed document), falls back to OCR page by page.
    """
    text_parts = []
    failed_pages = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            try:
                page_text = page.extract_text() or ""
            except Exception as exc:
                failed_pages.append(page_num)
                _logger.warning(f"page {page_num} of {pdf_path} failed to parse; skipping: {exc}")
                continue
            text_parts.append(page_text)

        used_ocr = False
        if not "".join(text_parts).strip():
            ocr_parts = [_ocr_page(page) for page in pdf.pages]
            ocr_parts = [t for t in ocr_parts if t.strip()]
            if ocr_parts:
                text_parts = ocr_parts
                used_ocr = True

    return "\n".join(text_parts), failed_pages, used_ocr


def decode_entities(tokens, pred_labels, confidences):
    """Turn a per-token BIO prediction into {entity: [values]} spans, plus
    the matching per-span confidence (the weakest token in a span, since a
    multi-token span is only as trustworthy as its shakiest token).

    Generic over whatever entity types the label set defines (B-Vendor,
    B-Merchant, B-Account, ...) rather than hardcoding one field, so it
    doesn't need to change as document kinds are added."""
    fields = {}
    confs = {}
    span_entity = None
    span_tokens, span_confs = [], []

    def flush_span():
        if span_entity and span_tokens:
            fields.setdefault(span_entity, []).append(" ".join(span_tokens))
            confs.setdefault(span_entity, []).append(min(span_confs))
        span_tokens.clear()
        span_confs.clear()

    for tok, label, conf in zip(tokens, pred_labels, confidences):
        if label.startswith("B-"):
            flush_span()
            span_entity = label[2:]
            span_tokens.append(tok)
            span_confs.append(conf)
        elif label.startswith("I-") and label[2:] == span_entity:
            span_tokens.append(tok)
            span_confs.append(conf)
        else:
            flush_span()
            span_entity = None
    flush_span()
    return fields, confs


def _empty_result(note):
    """A flagged, all-empty row. The reason for the failure is logged to the
    console for whoever is running this; the note shown to the end user
    stays in plain language, with no exception text or internals in it."""
    return {
        "Type": None,
        "Fields": [],
        "Error": True,
        "Notes": note,
        "NeedsReview": True,
    }


def _friendly_notes(low_conf_fields, missing_fields, failed_pages):
    parts = []
    if missing_fields:
        parts.append("Couldn't find: " + ", ".join(missing_fields))
    if low_conf_fields:
        parts.append("Worth a second look: " + ", ".join(low_conf_fields))
    if failed_pages:
        word = "page" if len(failed_pages) == 1 else "pages"
        parts.append(f"Part of this document (the {word} in question) wasn't clear enough to read")
    return " · ".join(parts)


def process_invoice(model, vocab, pdf_path, confidence_threshold=DEFAULT_CONFIDENCE_THRESHOLD):
    """Never raises for document-level problems; a failure becomes a
    flagged row (Error set, NeedsReview true) instead of an exception, so a
    bulk run keeps going through the rest of the batch."""
    try:
        text, failed_pages, used_ocr = extract_text(pdf_path)
    except Exception as exc:
        errors.report_exception(exc, stage="extract_text", pdf_path=pdf_path)
        return _empty_result("This file couldn't be opened. It may be damaged or not a valid PDF.")

    tokens = tokenize(text)
    if not tokens:
        note = "No readable text was found in this document."
        if failed_pages:
            note = "This document wasn't clear enough to read."
        return _empty_result(note)

    try:
        ids = vocab.encode(tokens)
        window_ids = build_context_windows(ids, model.window)
        pred_ids, confidences = model.predict_with_confidence(window_ids)
        pred_labels = [ID2LABEL[i] for i in pred_ids]
        entity_values, entity_confs = decode_entities(tokens, pred_labels, confidences)
    except Exception as exc:
        errors.report_exception(exc, stage="inference", pdf_path=pdf_path)
        return _empty_result("Something went wrong while reading this document. Please try again.")

    doc_type = infer_document_type(entity_confs)
    field_config = TYPE_FIELD_CONFIG[doc_type]

    fields = []
    low_conf_labels = []
    missing_labels = []
    for entity_key, display_label in field_config:
        values = entity_values.get(entity_key) or []
        confs = entity_confs.get(entity_key) or []
        value = values[0] if values else ""
        conf = confs[0] if confs else None
        fields.append({"key": entity_key, "label": display_label, "value": value, "confidence": conf})
        if value and conf is not None and conf < confidence_threshold:
            low_conf_labels.append(f"{display_label} (we're {conf:.0%} sure)")
        elif not value:
            missing_labels.append(display_label)

    notes = _friendly_notes(low_conf_labels, missing_labels, failed_pages)
    if used_ocr:
        ocr_note = "This document had no selectable text, so Nuru read it with OCR. Please double-check everything below."
        notes = f"{ocr_note} · {notes}" if notes else ocr_note

    return {
        "Type": doc_type,
        "Fields": fields,
        "Error": False,
        "Notes": notes,
        "NeedsReview": bool(low_conf_labels) or bool(failed_pages) or bool(missing_labels) or used_ocr,
    }


_MIN_COL_WIDTH = 10
_MAX_COL_WIDTH = 34
_DETAILS_COL_WIDTH = 46


def _style_header_cell(cell):
    cell.fill = HEADER_FILL
    cell.font = _font(size=11, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _CELL_BORDER


def _style_title_band(ws, total_cols, subtitle, meta):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value="NURU")
    title_cell.font = _font(size=22, bold=True, color=_NAVY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    subtitle_cell = ws.cell(row=2, column=1, value=subtitle)
    subtitle_cell.font = _font(size=12, color=_TEXT_SOFT)
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
    meta_cell = ws.cell(row=3, column=1, value=meta)
    meta_cell.font = _font(size=9, italic=True, color=_TEXT_SOFT)
    meta_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 16

    accent = ws.cell(row=4, column=1)
    for col in range(1, total_cols + 1):
        ws.cell(row=4, column=col).fill = HEADER_FILL
    ws.row_dimensions[4].height = 3


def write_excel(rows, output_path, confidence_threshold=DEFAULT_CONFIDENCE_THRESHOLD):
    """Each row's own inferred type picks its own field labels, so a batch
    mixing invoices, receipts, and statements doesn't force one header row
    onto all of them. Column layout is generic slots (Field 1/Value 1, ...),
    sized to whatever the widest row in this call actually has (a
    user-reviewed row can carry more custom fields than any built-in type
    schema), since a flat sheet can't otherwise represent rows with
    genuinely different schemas.

    Every column then auto-sizes to its own content (capped, with the
    notes column wrapping instead of growing unbounded) so nothing needs
    to be resized by hand after opening the file, and the report reads
    as a finished document rather than a raw data dump: a branded title
    band, colored status chips instead of plain text, real currency
    numbers where a value looks like money, and a short legend sheet
    explaining what each color means."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Records"
    ws.sheet_view.showGridLines = False

    max_fields = max((len(row.get("Fields", [])) for row in rows), default=0)
    max_fields = max(max_fields, 1)
    headers = ["Document", "Type"]
    for i in range(1, max_fields + 1):
        headers += [f"Field {i}", f"Value {i}"]
    headers += ["Please Verify", "Details"]
    total_cols = len(headers)

    generated = datetime.datetime.now().strftime("%B %d, %Y at %H:%M")
    doc_word = "document" if len(rows) == 1 else "documents"
    _style_title_band(
        ws, total_cols, "Document Extraction Report",
        f"Generated {generated}. {len(rows)} {doc_word}. "
        f"Fields below {confidence_threshold:.0%} confidence are flagged for review.",
    )

    header_row = 5
    for col, text in enumerate(headers, start=1):
        _style_header_cell(ws.cell(row=header_row, column=col, value=text))
    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{ws.cell(row=header_row, column=total_cols).coordinate}"

    details_col = total_cols
    verify_col = total_cols - 1

    for row in rows:
        fields = row.get("Fields", [])
        notes = row.get("Notes") or ""
        r = ws.max_row + 1

        doc_cell = ws.cell(row=r, column=1, value=row["File"])
        doc_cell.font = _font(bold=True)

        doc_type = row.get("Type")
        type_cell = ws.cell(row=r, column=2, value=doc_type or "Unrecognized")
        chip_text, chip_fill = _TYPE_CHIP.get(doc_type, (_TEXT_SOFT, "F1F2F6"))
        type_cell.font = _font(bold=True, color=chip_text)
        type_cell.fill = PatternFill(start_color=chip_fill, end_color=chip_fill, fill_type="solid")
        type_cell.alignment = Alignment(horizontal="center", vertical="center")

        for i in range(max_fields):
            label_col, value_col = 3 + i * 2, 4 + i * 2
            field = fields[i] if i < len(fields) else None
            label = field["label"] if field else ""
            value = field["value"] if field else ""

            label_cell = ws.cell(row=r, column=label_col, value=label)
            label_cell.font = _font(color=_TEXT_SOFT)

            value_cell = ws.cell(row=r, column=value_col)
            amount = _parse_money(value)
            if amount is not None:
                value_cell.value = amount
                value_cell.number_format = "$#,##0.00"
                value_cell.alignment = Alignment(horizontal="right")
            else:
                value_cell.value = value
            value_cell.font = _font()

            if field is not None:
                conf = field.get("confidence")
                low_confidence = value and conf is not None and conf < confidence_threshold
                missing = not value
                if low_confidence or missing:
                    value_cell.fill = LOW_CONFIDENCE_FILL
                    value_cell.font = _font(bold=True, color=_WARN_TEXT)
                    note = ("This field could not be found in the document."
                            if missing else
                            f"Nuru was only {conf:.0%} confident in this value. Please verify against the source document.")
                    value_cell.comment = Comment(note, "Nuru")

        is_error = bool(row.get("Error"))
        needs_review = bool(row.get("NeedsReview"))
        if is_error:
            verify_text = "Could not scan"
        elif needs_review:
            verify_text = "Please double-check"
        else:
            verify_text = "Looks good"
        verify_cell = ws.cell(row=r, column=verify_col, value=verify_text)
        verify_cell.alignment = Alignment(horizontal="center", vertical="center")
        if not is_error:
            if needs_review:
                verify_cell.fill = VERIFY_WARN_FILL
                verify_cell.font = _font(bold=True, color=_WARN_TEXT)
            else:
                verify_cell.fill = VERIFY_OK_FILL
                verify_cell.font = _font(bold=True, color=_OK_TEXT)

        details_cell = ws.cell(row=r, column=details_col, value=notes)
        details_cell.font = _font(color=_TEXT_SOFT)
        details_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = _wrapped_row_height(notes, _DETAILS_COL_WIDTH)

        if is_error:
            for col in range(1, total_cols + 1):
                cell = ws.cell(row=r, column=col)
                cell.fill = ERROR_FILL
                cell.font = _font(bold=(col == 1), color=_ERROR_TEXT)

        for col in range(1, total_cols + 1):
            ws.cell(row=r, column=col).border = _CELL_BORDER

    for col in range(1, total_cols + 1):
        letter = ws.cell(row=header_row, column=col).column_letter
        if col == details_col:
            ws.column_dimensions[letter].width = _DETAILS_COL_WIDTH
            continue
        header_len = len(headers[col - 1])
        content_len = max(
            (len(str(ws.cell(row=r, column=col).value))
             for r in range(header_row + 1, ws.max_row + 1)
             if ws.cell(row=r, column=col).value is not None),
            default=0,
        )
        width = max(header_len, content_len) + 3
        ws.column_dimensions[letter].width = min(max(width, _MIN_COL_WIDTH), _MAX_COL_WIDTH)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    _write_legend_sheet(wb, len(rows), confidence_threshold)
    wb.save(output_path)


def _write_legend_sheet(wb, document_count, confidence_threshold):
    ws = wb.create_sheet("About this report")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 70

    _style_title_band(
        ws, 2, "About this report",
        "What the colors in the Records sheet mean, and how this file was produced.",
    )

    rows = [
        ("Colored type chip", "The kind of document Nuru recognized: Invoice, Receipt, or Statement, "
                               "inferred from which fields were actually found, not fixed in advance."),
        ("Amber value cell", "Nuru flagged this specific value for a human to verify, either because "
                              "it was not found or because the model's own confidence in it was below "
                              f"{confidence_threshold:.0%}. Hover the cell for the exact reason."),
        ("Green Please Verify chip", "Every field on this row met the confidence threshold. Still worth "
                                      "a glance, not a guarantee."),
        ("Amber Please Verify chip", "At least one field on this row needs a second look before you "
                                      "rely on it."),
        ("Red row", "This document could not be read at all. See its Details cell for why."),
    ]
    header_row = 5
    for col, text in enumerate(["What you see", "What it means"], start=1):
        _style_header_cell(ws.cell(row=header_row, column=col, value=text))
    ws.row_dimensions[header_row].height = 20

    for i, (label, meaning) in enumerate(rows):
        r = header_row + 1 + i
        label_cell = ws.cell(row=r, column=1, value=label)
        label_cell.font = _font(bold=True)
        label_cell.alignment = Alignment(vertical="top", wrap_text=True)
        meaning_cell = ws.cell(row=r, column=2, value=meaning)
        meaning_cell.font = _font(color=_TEXT_SOFT)
        meaning_cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = _wrapped_row_height(meaning, 70)
        for col in (1, 2):
            ws.cell(row=r, column=col).border = _CELL_BORDER
        if i % 2 == 1:
            for col in (1, 2):
                ws.cell(row=r, column=col).fill = BAND_FILL

    footer_row = header_row + len(rows) + 2
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=2)
    footer_cell = ws.cell(
        row=footer_row, column=1,
        value=f"Covers {document_count} document(s). Extraction runs entirely on this machine, using a "
              f"hand-built model with no pretrained weights; no document content or extracted value is "
              f"sent to any third-party AI service.",
    )
    footer_cell.font = _font(size=9, italic=True, color=_TEXT_SOFT)
    footer_cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[footer_row].height = _wrapped_row_height(footer_cell.value, 90)


def main():
    parser = argparse.ArgumentParser(description="Nuru: read invoices, get organized records back.")
    parser.add_argument("pdfs", nargs="+", help="PDF file paths or glob patterns.")
    parser.add_argument("--output", default="nuru-invoices.xlsx")
    parser.add_argument("--confidence-threshold", type=float, default=DEFAULT_CONFIDENCE_THRESHOLD,
                         help="Softmax probability below which a field is flagged for review (default 0.70).")
    parser.add_argument("--purge-source", action="store_true",
                         help="Permanently delete each source PDF once it has been successfully "
                              "captured in the output Excel file. Off by default, since these are your "
                              "original files on disk, not a server-side cache; only enable this if "
                              "that is genuinely what you want.")
    args = parser.parse_args()

    if not os.path.exists(WEIGHTS_PATH) or not os.path.exists(VOCAB_PATH):
        raise SystemExit(
            f"No trained model found at {WEIGHTS_PATH} / {VOCAB_PATH}.\n"
            f"Run `python train.py` first."
        )

    vocab = Vocabulary.load(VOCAB_PATH)
    model = TokenClassifier.load_weights(WEIGHTS_PATH)

    pdf_paths = []
    for pattern in args.pdfs:
        matches = glob.glob(pattern)
        pdf_paths.extend(matches if matches else [pattern])

    rows = []
    succeeded_paths = []
    for pdf_path in pdf_paths:
        if not os.path.exists(pdf_path):
            print(f"skip (not found): {pdf_path}")
            continue
        print(f"processing: {pdf_path}")
        try:
            fields = process_invoice(model, vocab, pdf_path, args.confidence_threshold)
        except Exception as exc:
            # Belt-and-suspenders: process_invoice already catches document-level
            # failures, but nothing should be able to take the whole batch down.
            errors.report_exception(exc, stage="cli_batch", pdf_path=pdf_path)
            fields = _empty_result("Something went wrong while reading this document. Please try again.")
        fields["File"] = os.path.basename(pdf_path)
        rows.append(fields)
        if not fields.get("Error"):
            succeeded_paths.append(pdf_path)

    if not rows:
        raise SystemExit("No PDFs were processed.")

    write_excel(rows, args.output, args.confidence_threshold)
    print(f"Wrote {len(rows)} row(s) to {args.output}")

    if args.purge_source:
        for pdf_path in succeeded_paths:
            try:
                os.remove(pdf_path)
                print(f"purged source: {pdf_path}")
            except OSError as exc:
                _logger.warning(f"could not purge {pdf_path}: {exc}")


if __name__ == "__main__":
    main()
