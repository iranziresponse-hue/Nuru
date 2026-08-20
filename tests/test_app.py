import pytest
from openpyxl import load_workbook

from app import (
    TYPE_FIELD_CONFIG, VOCAB_PATH, WEIGHTS_PATH,
    _parse_money, _wrapped_row_height,
    decode_entities, infer_document_type, process_invoice, write_excel,
)


def test_infer_document_type_priority():
    assert infer_document_type({"Vendor": [0.95], "Tax": [0.9]}) == "Invoice"
    assert infer_document_type({"Merchant": [0.95]}) == "Receipt"
    assert infer_document_type({"PaymentMethod": [0.9]}) == "Receipt"
    assert infer_document_type({"Account": [0.9], "Balance": [0.9]}) == "Statement"
    assert infer_document_type({}) == "Invoice"  # fallback


def test_infer_document_type_ignores_low_confidence_noise():
    """A single near-coin-flip Merchant guess (e.g. from an <UNK> embedding
    on an out-of-vocabulary name) must not override strong, confident
    evidence for a different type elsewhere in the same document."""
    entity_confs = {
        "Vendor": [0.95], "Date": [1.0], "Total": [1.0], "Tax": [1.0],
        "Merchant": [0.45],  # spurious, low-confidence
    }
    assert infer_document_type(entity_confs) == "Invoice"


def test_infer_document_type_still_trusts_confident_receipt_signal():
    entity_confs = {"Merchant": [0.92], "PaymentMethod": [0.88]}
    assert infer_document_type(entity_confs) == "Receipt"


def test_infer_document_type_tax_beats_even_a_confident_merchant_guess():
    """The harder failure mode found by evaluate.py: an unfamiliar vendor
    name pattern-matched training's merchant names well enough to tag as
    Merchant at *high* confidence (0.95), while a correctly-detected Tax
    field (which never appears in receipt/statement training data, so
    finding one at all is close to unambiguous) sat right next to it. A
    confidence floor alone doesn't fix this; Tax has to be checked first."""
    entity_confs = {"Merchant": [0.95], "Tax": [1.0], "Date": [1.0], "Total": [1.0]}
    assert infer_document_type(entity_confs) == "Invoice"


def test_type_field_config_covers_every_document_kind():
    assert set(TYPE_FIELD_CONFIG.keys()) == {"Invoice", "Receipt", "Statement"}
    for type_name, fields in TYPE_FIELD_CONFIG.items():
        assert fields, f"{type_name} has no configured fields"
        keys = [k for k, _ in fields]
        assert len(keys) == len(set(keys)), f"{type_name} has duplicate entity keys"


def test_decode_entities_single_token_span():
    tokens = ["Total", ":", "$500.00"]
    labels = ["O", "O", "B-Total"]
    confs = [0.9, 0.9, 0.95]
    values, confidences = decode_entities(tokens, labels, confs)
    assert values == {"Total": ["$500.00"]}
    assert confidences == {"Total": [0.95]}


def test_decode_entities_multi_token_span_uses_min_confidence():
    tokens = ["Acme", "Supply", "Co", "Invoice"]
    labels = ["B-Vendor", "I-Vendor", "I-Vendor", "O"]
    confs = [0.99, 0.80, 0.95, 0.5]
    values, confidences = decode_entities(tokens, labels, confs)
    assert values == {"Vendor": ["Acme Supply Co"]}
    assert confidences["Vendor"] == [0.80]  # weakest token in the span


def test_decode_entities_orphan_continuation_tag_does_not_attach_to_stale_span():
    # A B-Vendor span is flushed by an O, then a later I-Vendor with no
    # matching B- shouldn't silently reattach to the old span's leftovers.
    tokens = ["Acme", "and", "Co"]
    labels = ["B-Vendor", "O", "I-Vendor"]
    confs = [0.9, 0.9, 0.9]
    values, _ = decode_entities(tokens, labels, confs)
    assert values == {"Vendor": ["Acme"]}  # "Co" must NOT be silently appended


def test_decode_entities_no_entities_found():
    values, confidences = decode_entities(["hello", "world"], ["O", "O"], [0.9, 0.9])
    assert values == {}
    assert confidences == {}


# ---- end-to-end extraction against a real generated PDF, using the
# repo's actual trained weights (checked in) ----------------------------

reportlab_canvas = pytest.importorskip("reportlab.pdfgen.canvas")


def _make_pdf(path, lines):
    c = reportlab_canvas.Canvas(str(path))
    y = 750
    for line in lines:
        c.drawString(72, y, line)
        y -= 20
    c.save()


@pytest.fixture(scope="module")
def model_and_vocab():
    import os
    if not os.path.exists(WEIGHTS_PATH) or not os.path.exists(VOCAB_PATH):
        pytest.skip("model not trained; run train.py first")
    from engine.model import TokenClassifier
    from engine.tokenizer import Vocabulary
    return TokenClassifier.load_weights(WEIGHTS_PATH), Vocabulary.load(VOCAB_PATH)


def test_process_invoice_extracts_invoice_fields(tmp_path, model_and_vocab):
    model, vocab = model_and_vocab
    pdf_path = tmp_path / "invoice.pdf"
    _make_pdf(pdf_path, [
        "Quantum Electronics", "Invoice Number: INV-88213", "Invoice Date: 2026-04-02",
        "Bill To: Customer Account # 5521", "Subtotal: $1,240.00", "Tax (8.0%): $99.20",
        "Total Due: $1,339.20", "Payment is due within 30 days. Thank you for your business.",
    ])
    result = process_invoice(model, vocab, str(pdf_path))
    assert result["Error"] is False
    assert result["Type"] == "Invoice"
    values = {f["label"]: f["value"] for f in result["Fields"]}
    assert values["Supplier Name"] == "Quantum Electronics"
    assert values["Total Amount Due"] == "$1,339.20"


def test_process_invoice_handles_corrupt_pdf_without_raising(tmp_path, model_and_vocab):
    model, vocab = model_and_vocab
    bad_path = tmp_path / "not_a_pdf.pdf"
    bad_path.write_bytes(b"this is not a real pdf")
    result = process_invoice(model, vocab, str(bad_path))
    assert result["Error"] is True
    assert result["Fields"] == []
    assert result["NeedsReview"] is True


def test_held_out_evaluation_set_meets_a_regression_floor(model_and_vocab):
    """Runs the actual eval/documents/*.pdf held-out set (deliberately
    novel names/phrasing absent from training; see eval/generate_eval_set.py)
    through the real evaluation harness. Not a claim about real-world
    accuracy (these are still synthetic documents), just a floor that
    catches a future change quietly regressing extraction quality."""
    import json
    import os

    from evaluate import evaluate_document, summarize

    model, vocab = model_and_vocab
    ground_truth_path = os.path.join("eval", "ground_truth.json")
    documents_dir = os.path.join("eval", "documents")
    if not os.path.exists(ground_truth_path):
        pytest.skip("eval set not generated; run eval/generate_eval_set.py first")

    with open(ground_truth_path, "r", encoding="utf-8") as f:
        ground_truth = json.load(f)

    all_results, type_matches = [], 0
    for filename, expected in ground_truth.items():
        results, inferred_type = evaluate_document(
            model, vocab, os.path.join(documents_dir, filename), expected, 0.70,
        )
        all_results.extend(results)
        if inferred_type == expected["type"]:
            type_matches += 1

    summary = summarize(all_results, type_matches, len(ground_truth))
    assert summary["document_type_accuracy"] >= 0.90
    assert summary["field_accuracy"] >= 0.85


# ---- write_excel ------------------------------------------------------------

def test_parse_money_recognizes_currency_strings():
    assert _parse_money("$1,339.20") == 1339.20
    assert _parse_money("(200.00)") == -200.0
    assert _parse_money("47.88") == 47.88


def test_parse_money_rejects_non_money_text():
    assert _parse_money("Acme Co") is None
    assert _parse_money("") is None
    assert _parse_money(None) is None


def test_wrapped_row_height_grows_with_text_length():
    short = _wrapped_row_height("hi", width_chars=40)
    long = _wrapped_row_height("x" * 200, width_chars=40)
    assert long > short
    assert _wrapped_row_height("", width_chars=40) == 18


def _sample_rows():
    return [
        {
            "File": "invoice_acme.pdf", "Type": "Invoice",
            "Fields": [
                {"label": "Supplier Name", "value": "Acme Co", "confidence": 0.95, "key": "Vendor"},
                {"label": "Transaction Date", "value": "2026-01-15", "confidence": 0.4, "key": "Date"},
                {"label": "Total Amount Due", "value": "$1,339.20", "confidence": 0.99, "key": "Total"},
                {"label": "Tax Value (VAT/GST)", "value": "", "confidence": None, "key": "Tax"},
            ],
            "NeedsReview": True, "Notes": "Needs a second look.", "Error": False,
        },
        {
            "File": "receipt_shop.pdf", "Type": "Receipt",
            "Fields": [
                {"label": "Merchant", "value": "Copperline Hardware", "confidence": 0.9, "key": "Merchant"},
            ],
            "NeedsReview": False, "Notes": "", "Error": False,
        },
        {
            "File": "broken.pdf", "Type": None, "Fields": [],
            "NeedsReview": False, "Notes": "Could not be read.", "Error": True,
        },
    ]


def test_write_excel_produces_a_records_and_legend_sheet(tmp_path):
    out = tmp_path / "report.xlsx"
    write_excel(_sample_rows(), str(out), confidence_threshold=0.70)
    wb = load_workbook(str(out))
    assert wb.sheetnames == ["Records", "About this report"]


def test_write_excel_money_like_values_become_real_numbers(tmp_path):
    out = tmp_path / "report.xlsx"
    write_excel(_sample_rows(), str(out), confidence_threshold=0.70)
    ws = load_workbook(str(out))["Records"]
    # Row 6 = first data row (rows 1-4 are the title band, row 5 is headers).
    # Field 3 / Value 3 = Total Amount Due, at columns 7/8.
    assert ws.cell(row=6, column=8).value == 1339.20
    assert ws.cell(row=6, column=8).number_format == "$#,##0.00"


def test_write_excel_flags_low_confidence_and_missing_values_with_a_note(tmp_path):
    out = tmp_path / "report.xlsx"
    write_excel(_sample_rows(), str(out), confidence_threshold=0.70)
    ws = load_workbook(str(out))["Records"]
    # Field 2 / Value 2 = Transaction Date (confidence 0.4, below threshold), columns 5/6.
    low_conf_cell = ws.cell(row=6, column=6)
    assert low_conf_cell.comment is not None
    assert "confident" in low_conf_cell.comment.text
    # Field 4 / Value 4 = Tax (missing), columns 9/10.
    missing_cell = ws.cell(row=6, column=10)
    assert missing_cell.comment is not None
    assert "could not be found" in missing_cell.comment.text


def test_write_excel_verify_column_reflects_row_state(tmp_path):
    out = tmp_path / "report.xlsx"
    write_excel(_sample_rows(), str(out), confidence_threshold=0.70)
    ws = load_workbook(str(out))["Records"]
    assert ws.cell(row=6, column=11).value == "Please double-check"  # needs review
    assert ws.cell(row=7, column=11).value == "Looks good"           # clean
    assert ws.cell(row=8, column=11).value == "Could not scan"       # error, not "Looks good"


def test_write_excel_error_row_is_not_falsely_marked_looks_good(tmp_path):
    """A document that failed to scan must never claim to look good, even
    though NeedsReview defaults to False for it same as a clean row."""
    out = tmp_path / "report.xlsx"
    write_excel(_sample_rows(), str(out), confidence_threshold=0.70)
    ws = load_workbook(str(out))["Records"]
    assert ws.cell(row=8, column=11).value != "Looks good"


def test_write_excel_columns_are_auto_sized_within_bounds(tmp_path):
    out = tmp_path / "report.xlsx"
    write_excel(_sample_rows(), str(out), confidence_threshold=0.70)
    ws = load_workbook(str(out))["Records"]
    for letter, dim in ws.column_dimensions.items():
        if dim.width is not None:
            assert 10 <= dim.width <= 46


def test_write_excel_no_em_dashes_or_curly_quotes_anywhere(tmp_path):
    out = tmp_path / "report.xlsx"
    write_excel(_sample_rows(), str(out), confidence_threshold=0.70)
    wb = load_workbook(str(out))
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    assert "—" not in cell.value
                if cell.comment:
                    assert "—" not in cell.comment.text
